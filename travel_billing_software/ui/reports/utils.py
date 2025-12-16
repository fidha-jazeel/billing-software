"""
Utility Functions for Reports Module
Provides table configuration, filtering, export, and UI helpers.
"""
from datetime import datetime
from typing import List, Dict, Any, Optional
from PyQt6.QtWidgets import (
    QTableWidget, QFrame, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QDateEdit, QComboBox, QPushButton, QHeaderView,
    QMessageBox, QFileDialog, QWidget
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QColor
from travel_billing_software.utils.logger import log_info, log_error, log_warning
from travel_billing_software.config.config import format_currency, get_currency_symbol
from travel_billing_software.database.db_manager import get_db_instance



class TableConfigurator:
    """
    Handles table widget configuration with consistent styling.
    
    Provides methods for:
    - Setting up sortable headers
    - Configuring column widths
    - Applying consistent visual styles
    """
    
    @staticmethod
    def configure_table(table: QTableWidget, column_widths: Optional[Dict[int, Any]] = None):
        """
        Configure table with proper sizing, sorting, and styling.
        
        Args:
            table: QTableWidget to configure
            column_widths: Dict mapping column index to width (None for auto-resize)
                          Use 'stretch' for QHeaderView.Stretch mode
                          Use int for fixed pixel width
        """
        # Enable sorting
        table.setSortingEnabled(True)
        
        # Configure header
        header = table.horizontalHeader()
        header.setStyleSheet("""
            QHeaderView::section {
                background-color: #000000;
                color: #FFFFFF;
                padding: 12px 8px;
                border: 1px solid #777777;
                border-bottom: 1px solid #777777;
                font-weight: 600;
                font-size: 15px;
                text-align: left;
            }
            QHeaderView::section:hover {
                background-color: #222222;
            }
        """)
        header.setMinimumHeight(35)
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        
        # Make header sticky and non-movable
        header.setStretchLastSection(False)
        header.setSectionsMovable(False)
        header.setHighlightSections(True)
        
        # Set column widths if provided
        if column_widths:
            for col, width in column_widths.items():
                if width == 'stretch':
                    header.setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)
                elif width == 'auto':
                    # Auto-resize based on content
                    header.setSectionResizeMode(col, QHeaderView.ResizeMode.ResizeToContents)
                elif isinstance(width, int):
                    table.setColumnWidth(col, width)
                    # Use Interactive mode to allow manual resizing while respecting minimum width
                    header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
        
        # Enable text wrapping for better display of long content
        table.setWordWrap(True)
        
        log_info(f"Table configured with {table.columnCount()} columns", 'billing_app')


class ReportFilters:
    """
    Manages filter creation and application for reports.
    
    Provides unified filtering UI components and logic for:
    - Date range filtering
    - Contact/customer search
    - Passenger name search
    - Sector filtering
    - Supplier filtering
    - Booking type filtering
    """
    
    def __init__(self, colors: dict, get_button_style: callable, db_manager=None):
        """
        Initialize filter manager.
        
        Args:
            colors: Color scheme dictionary
            get_button_style: Function to get button stylesheet
            db_manager: Database manager instance (optional, will use singleton if not provided)
        """
        self.colors = colors
        self.get_button_style = get_button_style
        self.db = db_manager if db_manager else get_db_instance()
        
        # Filter widgets - will be created in create_filter_section()
        self.filter_from_date = None
        self.filter_to_date = None
        self.filter_contact = None
        self.filter_passenger = None
        self.filter_customer_name = None
        self.filter_sector = None
        self.filter_supplier = None
        self.filter_type = None
    
    def load_suppliers_from_db(self) -> List[str]:
        """
        Load supplier names from database and return as sorted list.
        
        Returns:
            List of supplier names sorted alphabetically, with 'All' as first item
        """
        try:
            # Get all contacts with type 'SUPPLIER'
            suppliers = self.db.get_contacts(contact_type='SUPPLIER')
            
            # Extract supplier names
            supplier_names = [s.get('name', '') for s in suppliers if s.get('name')]
            
            # Sort alphabetically
            supplier_names.sort()
            
            # Add 'All' at the beginning
            supplier_list = ['All'] + supplier_names
            
            log_info(f"Loaded {len(supplier_names)} suppliers from database", 'billing_app')
            return supplier_list
            
        except Exception as e:
            log_error("Failed to load suppliers from database", exception=e, logger_name='billing_errors')
            # Return default list as fallback
            return ['All', 'IndiGo', 'Air India', 'SpiceJet', 'Vistara', 'AirAsia', 'Other']
    
    def load_types_from_db(self) -> List[str]:
        """
        Load booking/invoice types from database and return as sorted list.
        
        Returns:
            List of type names sorted alphabetically, with 'All' as first item
        """
        try:
            # Get all types from dropdown_types table
            types = self.db.get_dropdown_items('type')
            
            # Sort alphabetically
            types.sort()
            
            # Add 'All' at the beginning
            type_list = ['All'] + types
            
            log_info(f"Loaded {len(types)} types from database", 'billing_app')
            return type_list
            
        except Exception as e:
            log_error("Failed to load types from database", exception=e, logger_name='billing_errors')
            # Return default list as fallback
            return ['All', 'Flight', 'Hotel', 'Visa', 'Tour Package', 'Insurance', 'Other']
    
    def create_filter_section(self, apply_callback: callable, clear_callback: callable) -> QFrame:
        """
        Create comprehensive filter section with all controls (collapsible).
        
        Args:
            apply_callback: Function to call when Apply button clicked
            clear_callback: Function to call when Clear button clicked
            
        Returns:
            QFrame containing all filter controls
        """
        filter_frame = QFrame()
        filter_frame.setStyleSheet("""
            QFrame {
                background-color: #0F0F0F;
                border-radius: 8px;
                border: 2px solid #777777;
                padding: 0px;
                margin: 0px;
            }
        """)
        
        filter_layout = QVBoxLayout(filter_frame)
        filter_layout.setSpacing(10)
        filter_layout.setContentsMargins(14, 14, 14, 14)
        
        # Title with toggle button
        title_layout = QHBoxLayout()
        filter_title = QLabel("🔍 Filter Options")
        filter_title.setStyleSheet("""
            QLabel {
                color: #FFFFFF;
                font-size: 14px;
                font-weight: bold;
                letter-spacing: 0.5px;
                padding: 8px;
                margin: 0px;
                border: none;
                background-color: transparent;
            }
        """)
        title_layout.addWidget(filter_title)
        
        toggle_btn = QPushButton("▼ Expand")
        toggle_btn.setStyleSheet("""
            QPushButton {
                color: #FFFFFF;
                background-color: #1A1A1A;
                border: 1px solid #777777;
                border-radius: 4px;
                padding: 6px 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2A2A2A;
            }
        """)
        toggle_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        title_layout.addWidget(toggle_btn)
        
        title_container = QFrame()
        title_container.setStyleSheet("""
            QFrame {
                background-color: #1A1A1A;
                border: 1px solid #777777;
                border-radius: 4px;
                padding: 0px;
                margin: 0px 0px 10px 0px;
            }
        """)
        title_container.setLayout(title_layout)
        filter_layout.addWidget(title_container)
        
        # Collapsible content widget
        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        content_layout.setSpacing(10)
        content_layout.setContentsMargins(0, 0, 0, 0)
        
        # Date Range
        date_row = QHBoxLayout()
        date_row.setSpacing(10)
        
        from_label = QLabel("From:")
        from_label.setStyleSheet("color: #FFFFFF; font-weight: bold;")
        date_row.addWidget(from_label)
        
        self.filter_from_date = QDateEdit()
        self.filter_from_date.setCalendarPopup(True)
        # Set to year 1900 to include ALL historical invoices by default
        # No automatic date filtering - show everything unless user manually changes
        self.filter_from_date.setDate(QDate(1900, 1, 1))
        self.filter_from_date.setStyleSheet(self._get_dateedit_style())
        date_row.addWidget(self.filter_from_date)
        
        to_label = QLabel("To:")
        to_label.setStyleSheet("color: #FFFFFF; font-weight: bold;")
        date_row.addWidget(to_label)
        
        self.filter_to_date = QDateEdit()
        self.filter_to_date.setCalendarPopup(True)
        # Set to 100 years in future to include all future-dated invoices
        self.filter_to_date.setDate(QDate.currentDate().addYears(100))
        self.filter_to_date.setStyleSheet(self._get_dateedit_style())
        date_row.addWidget(self.filter_to_date)
        
        content_layout.addLayout(date_row)
        
        # Contact Number
        contact_row = self._create_input_row("Contact:", "Search by contact number...")
        self.filter_contact = contact_row['input']
        content_layout.addLayout(contact_row['layout'])
        
        # Customer Name
        customer_row = self._create_input_row("Customer Name:", "Search by customer name...")
        self.filter_customer_name = customer_row['input']
        content_layout.addLayout(customer_row['layout'])

        # Passenger Name
        passenger_row = self._create_input_row("Passenger:", "Search by passenger name...")
        self.filter_passenger = passenger_row['input']
        content_layout.addLayout(passenger_row['layout'])
        
        # Sector and Supplier
        sector_supplier_row = QHBoxLayout()
        sector_supplier_row.setSpacing(10)
        
        sector_section = self._create_input_row("Sector:", "Search by sector...")
        self.filter_sector = sector_section['input']
        sector_supplier_row.addLayout(sector_section['layout'])
        
        # Load suppliers dynamically from database
        supplier_list = self.load_suppliers_from_db()
        supplier_section = self._create_combo_row(
            "Supplier:",
            supplier_list
        )
        self.filter_supplier = supplier_section['combo']
        sector_supplier_row.addLayout(supplier_section['layout'])
        
        content_layout.addLayout(sector_supplier_row)
        
        # Booking Type - Load dynamically from database
        type_list = self.load_types_from_db()
        type_section = self._create_combo_row(
            "Type:",
            type_list
        )
        self.filter_type = type_section['combo']
        content_layout.addLayout(type_section['layout'])
        
        # Apply and Clear buttons
        btn_row = QHBoxLayout()
        
        apply_btn = QPushButton("✓ Apply Filters")
        apply_btn.setStyleSheet(self.get_button_style('add'))
        apply_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        apply_btn.clicked.connect(apply_callback)
        btn_row.addWidget(apply_btn)
        
        clear_btn = QPushButton("✕ Clear")
        clear_btn.setStyleSheet(self.get_button_style('delete'))
        clear_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        clear_btn.clicked.connect(clear_callback)
        btn_row.addWidget(clear_btn)
        
        content_layout.addLayout(btn_row)
        
        # Initially hide content (collapsed state)
        content_widget.setVisible(False)
        filter_layout.addWidget(content_widget)
        
        # Connect toggle button
        def toggle_filters():
            is_visible = content_widget.isVisible()
            content_widget.setVisible(not is_visible)
            toggle_btn.setText("▲ Collapse" if not is_visible else "▼ Expand")
        
        toggle_btn.clicked.connect(toggle_filters)
        
        # Store widget references as properties of the filter_frame for later retrieval
        filter_frame.filter_from_date = self.filter_from_date
        filter_frame.filter_to_date = self.filter_to_date
        filter_frame.filter_contact = self.filter_contact
        filter_frame.filter_passenger = self.filter_passenger
        filter_frame.filter_customer_name = self.filter_customer_name
        filter_frame.filter_sector = self.filter_sector
        filter_frame.filter_supplier = self.filter_supplier
        filter_frame.filter_type = self.filter_type
        
        log_info("Filter section created (collapsible)", 'billing_app')
        return filter_frame
    
    def _create_input_row(self, label_text: str, placeholder: str) -> Dict[str, Any]:
        """Create bordered label and input field row."""
        row_layout = QHBoxLayout()
        row_layout.setSpacing(10)
        
        # Label box
        label_box = QFrame()
        label_box.setStyleSheet("""
            QFrame {
                background-color: #121212;
                border: 1px solid #777777;
                border-radius: 5px;
                padding: 6px 10px;
            }
        """)
        label_layout = QHBoxLayout(label_box)
        label_layout.setContentsMargins(0, 0, 0, 0)
        label = QLabel(label_text)
        label.setStyleSheet("color: #FFFFFF; font-weight: bold; background: transparent; border: none;")
        label_layout.addWidget(label)
        row_layout.addWidget(label_box)
        
        # Input box
        input_box = QFrame()
        input_box.setStyleSheet("""
            QFrame {
                background-color: #000000;
                border: 1px solid #777777;
                border-radius: 5px;
                padding: 6px 10px;
            }
        """)
        input_layout = QHBoxLayout(input_box)
        input_layout.setContentsMargins(0, 0, 0, 0)
        
        input_field = QLineEdit()
        input_field.setPlaceholderText(placeholder)
        input_field.setStyleSheet("""
            QLineEdit {
                background-color: transparent;
                color: #FFFFFF;
                border: none;
                padding: 2px;
            }
            QLineEdit::placeholder {
                color: #CCCCCC;
            }
            QLineEdit:focus {
                outline: none;
                border: none;
            }
        """)
        input_layout.addWidget(input_field)
        row_layout.addWidget(input_box, 1)
        
        return {'layout': row_layout, 'input': input_field}
    
    def _create_combo_row(self, label_text: str, items: List[str]) -> Dict[str, Any]:
        """Create bordered label and combobox row."""
        row_layout = QHBoxLayout()
        row_layout.setSpacing(10)
        
        # Label box
        label_box = QFrame()
        label_box.setStyleSheet("""
            QFrame {
                background-color: #121212;
                border: 1px solid #777777;
                border-radius: 5px;
                padding: 6px 10px;
            }
        """)
        label_layout = QHBoxLayout(label_box)
        label_layout.setContentsMargins(0, 0, 0, 0)
        label = QLabel(label_text)
        label.setStyleSheet("color: #FFFFFF; font-weight: bold; background: transparent; border: none;")
        label_layout.addWidget(label)
        row_layout.addWidget(label_box)
        
        # Combo box
        combo_frame = QFrame()
        combo_frame.setStyleSheet("""
            QFrame {
                background-color: #000000;
                border: 1px solid #777777;
                border-radius: 5px;
                padding: 6px 10px;
            }
        """)
        combo_layout = QHBoxLayout(combo_frame)
        combo_layout.setContentsMargins(0, 0, 0, 0)
        
        combo = QComboBox()
        combo.addItems(items)
        combo.setStyleSheet("""
            QComboBox {
                background-color: transparent;
                color: #FFFFFF;
                border: none;
                padding: 2px;
            }
            QComboBox::drop-down {
                border: none;
                background-color: transparent;
            }
            QComboBox QAbstractItemView {
                background-color: #1A1A1A;
                color: #FFFFFF;
                selection-background-color: #333333;
                border: 1px solid #777777;
            }
            QComboBox:focus {
                outline: none;
                border: none;
            }
        """)
        combo_layout.addWidget(combo)
        row_layout.addWidget(combo_frame, 1)
        
        return {'layout': row_layout, 'combo': combo}
    
    def _get_dateedit_style(self) -> str:
        """Get stylesheet for QDateEdit widgets."""
        return """
            QDateEdit {
                background-color: #1A1A1A;
                color: #FFFFFF;
                border: none;
                border-radius: 4px;
                padding: 8px;
            }
            QDateEdit::drop-down {
                border: none;
                background-color: #1A1A1A;
            }
        """
    
    def apply_filters(self, invoices: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Apply current filter values to invoice list.
        
        Args:
            invoices: List of invoice dictionaries
            
        Returns:
            Filtered list of invoices
        """
        try:
            filtered = []

            from_date = self.filter_from_date.date().toPyDate()
            to_date = self.filter_to_date.date().toPyDate()
            contact = self.filter_contact.text().lower()
            passenger = self.filter_passenger.text().lower()
            customer_name = self.filter_customer_name.text().lower() if self.filter_customer_name else ""
            sector = self.filter_sector.text().lower()
            supplier = self.filter_supplier.currentText()
            booking_type = self.filter_type.currentText()

            log_info(
                f"Applying filters - Date: {from_date} to {to_date}, "
                f"Contact: '{contact}', Passenger: '{passenger}', Customer: '{customer_name}', Sector: '{sector}', "
                f"Supplier: '{supplier}', Type: '{booking_type}'",
                'billing_app'
            )

            for invoice in invoices:
                # Customer Name filter
                if customer_name and customer_name not in invoice.get('customer_name', '').lower():
                    continue
                try:
                    # Date filter - OPTIONAL, only applied if invoice has valid date
                    # Default behavior: ALWAYS INCLUDE invoices with invalid/missing dates
                    # This ensures 100% of invoices are visible by default
                    apply_date_filter = False
                    invoice_date = None
                    
                    try:
                        date_str = invoice.get('invoice_date') or invoice.get('date') or invoice.get('created_at', '')
                        if date_str and str(date_str).strip() and str(date_str) != 'None':
                            # Parse date from various formats
                            if '/' in str(date_str):
                                # Format: "07/12/2024" or "7/12/2024"
                                parts = str(date_str).split('/')
                                day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
                                invoice_date = datetime(year, month, day).date()
                                apply_date_filter = True
                            elif ' ' in str(date_str):
                                # Format: "2024-12-07 12:16:16"
                                date_part = str(date_str).split()[0]
                                invoice_date = datetime.strptime(date_part, '%Y-%m-%d').date()
                                apply_date_filter = True
                            else:
                                # Format: "2024-12-07"
                                invoice_date = datetime.strptime(str(date_str), '%Y-%m-%d').date()
                                apply_date_filter = True
                    except Exception as date_error:
                        # Date parsing failed - ALWAYS INCLUDE this invoice
                        # Date filter is optional, not mandatory
                        pass
                    
                    # Only apply date filter if we successfully parsed a valid date
                    # Otherwise, INCLUDE the invoice (no date filtering)
                    if apply_date_filter and invoice_date:
                        if not (from_date <= invoice_date <= to_date):
                            continue
                    
                    # Contact filter
                    if contact and contact not in invoice.get('customer_phone', '').lower():
                        continue
                    
                    # Booking type filter
                    if booking_type != "All":
                        tickets = invoice.get('tickets', [])
                        if not any(ticket.get('booking_type', '') == booking_type for ticket in tickets):
                            continue
                    
                    # Passenger, sector, supplier filters
                    if passenger or sector or supplier != "All":
                        match_found = False
                        tickets = invoice.get('tickets', [])
                        passengers_list = invoice.get('passengers', [])
                        
                        for ticket in tickets:
                            if passenger:
                                for pax in passengers_list:
                                    if passenger in pax.get('name', '').lower():
                                        match_found = True
                                        break
                            if sector and sector in ticket.get('sector', '').lower():
                                match_found = True
                            if supplier != "All" and supplier == ticket.get('supplier_name', ''):
                                match_found = True
                            if match_found:
                                break
                        
                        if not match_found and (passenger or sector or supplier != "All"):
                            continue
                    
                    filtered.append(invoice)
                    
                except Exception as invoice_error:
                    log_error(
                        f"Error filtering invoice {invoice.get('invoice_number', 'Unknown')}",
                        exception=invoice_error,
                        logger_name='billing_errors'
                    )
                    continue
            
            log_info(f"Filter applied - {len(filtered)} records matched out of {len(invoices)}", 'billing_app')
            return filtered
            
        except Exception as e:
            log_error("Error applying filters", exception=e, logger_name='billing_errors')
            return []
    
    def clear_filters(self):
        """Reset all filter values to defaults - show ALL invoices."""
        try:
            log_info("Clearing all filters - resetting to show ALL invoices", 'billing_app')
            
            # Reset date range to show ALL invoices (1900 to 100 years future)
            self.filter_from_date.setDate(QDate(1900, 1, 1))
            self.filter_to_date.setDate(QDate.currentDate().addYears(100))
            self.filter_contact.clear()
            self.filter_passenger.clear()
            self.filter_sector.clear()
            self.filter_supplier.setCurrentIndex(0)
            self.filter_type.setCurrentIndex(0)
            
            log_info("Filters cleared - now showing ALL invoices", 'billing_app')
            
        except Exception as e:
            log_error("Error clearing filters", exception=e, logger_name='billing_errors')
            raise
    
    def refresh_supplier_dropdown(self):
        """
        Refresh the supplier dropdown with latest data from database.
        This should be called when suppliers are added/updated in the Supplier page.
        """
        try:
            if self.filter_supplier is None:
                log_warning("Supplier dropdown not initialized yet", 'billing_app')
                return
            
            # Store current selection
            current_selection = self.filter_supplier.currentText()
            
            # Clear existing items
            self.filter_supplier.clear()
            
            # Reload suppliers from database
            supplier_list = self.load_suppliers_from_db()
            
            # Add updated items
            self.filter_supplier.addItems(supplier_list)
            
            # Restore previous selection if it still exists
            index = self.filter_supplier.findText(current_selection)
            if index >= 0:
                self.filter_supplier.setCurrentIndex(index)
            else:
                self.filter_supplier.setCurrentIndex(0)  # Default to 'All'
            
            log_info(f"Supplier dropdown refreshed with {len(supplier_list)} items", 'billing_app')
            
        except Exception as e:
            log_error("Error refreshing supplier dropdown", exception=e, logger_name='billing_errors')
    
    def refresh_type_dropdown(self):
        """
        Refresh the type dropdown with latest data from database.
        This should be called when types are added/updated in Settings → Types page.
        """
        try:
            if self.filter_type is None:
                log_warning("Type dropdown not initialized yet", 'billing_app')
                return
            
            # Store current selection
            current_selection = self.filter_type.currentText()
            
            # Clear existing items
            self.filter_type.clear()
            
            # Reload types from database
            type_list = self.load_types_from_db()
            
            # Add updated items
            self.filter_type.addItems(type_list)
            
            # Restore previous selection if it still exists
            index = self.filter_type.findText(current_selection)
            if index >= 0:
                self.filter_type.setCurrentIndex(index)
            else:
                self.filter_type.setCurrentIndex(0)  # Default to 'All'
            
            log_info(f"Type dropdown refreshed with {len(type_list)} items", 'billing_app')
            
        except Exception as e:
            log_error("Error refreshing type dropdown", exception=e, logger_name='billing_errors')


class ReportExporter:
    @staticmethod
    def _map_invoice_for_print(invoice: dict) -> dict:
        """
        Map the report invoice dict to the structure expected by generate_invoice_pdf.
        """
        if not invoice:
            return None
        # Company info (customize as needed or fetch from config)
        company = {
            'name': invoice.get('company_name', 'Company Name'),
            'address': invoice.get('company_address', ''),
            'footer_note': invoice.get('company_footer', ''),
            'logo_path': invoice.get('company_logo', ''),
        }
        # Invoice meta
        invoice_meta = {
            'number': invoice.get('invoice_number', ''),
            'date': invoice.get('invoice_date', ''),
            'customer_id': invoice.get('customer_id', ''),
        }
        # Customer
        customer = {
            'name': invoice.get('customer_name', ''),
            'address': invoice.get('customer_address', ''),
            'contact': invoice.get('customer_phone', ''),
        }
        # Items (tickets)
        items = []
        for idx, ticket in enumerate(invoice.get('tickets', []), 1):
            # Try to get the correct unit price and amount fields
            unit_price = (
                ticket.get('unit_price')
                or ticket.get('cust_amt')
                or ticket.get('customer_amount')
                or ticket.get('total_amount')
                or 0
            )
            amount = (
                ticket.get('total_amount')
                or ticket.get('amount')
                or ticket.get('cust_amt')
                or ticket.get('customer_amount')
                or unit_price
                or 0
            )
            items.append({
                'sno': idx,
                'passenger_name': ticket.get('passenger_name', ''),
                'pnr': ticket.get('pnr', ''),
                'sector': ticket.get('sector', ''),
                'type': ticket.get('booking_type', ''),
                'qty': ticket.get('quantity', 1),
                'unit_price': unit_price,
                'amount': amount,
            })
        # Totals
        subtotal = sum(float(t.get('total_amount', 0) or 0) for t in invoice.get('tickets', []))
        tax = float(invoice.get('tax', 0) or 0)
        total = subtotal + tax
        # Currency
        currency = invoice.get('currency', 'AED')
        # Discount, notes, terms
        discount = float(invoice.get('discount', 0) or 0)
        notes = invoice.get('notes', '')
        terms = invoice.get('terms', '')
        return {
            'company': company,
            'invoice_meta': invoice_meta,
            'customer': customer,
            'items': items,
            'discount': discount,
            'notes': notes,
            'terms': terms,
            'currency': currency,
            'tax': tax,
            'subtotal': subtotal,
            'total': total,
            'paid_amount': float(invoice.get('paid_amount', 0)),
            'balance': float(invoice.get('balance', total - float(invoice.get('paid_amount', 0))))
        }

    @staticmethod
    def print_invoice(invoice: dict, parent: QWidget):
        """
        Generate a temporary PDF for the given invoice and open it for printing or manual print if no association exists.
        """
        import tempfile
        import os
        import sys
        from PyQt6.QtWidgets import QMessageBox
        try:
            from travel_billing_software.utils.invoice_generator import generate_invoice_pdf
        except ImportError as e:
            log_error("Could not import generate_invoice_pdf", exception=e, logger_name='billing_errors')
            QMessageBox.critical(parent, "Print Error", "Could not import invoice PDF generator.")
            return

        # Map invoice to print template structure
        mapped_invoice = ReportExporter._map_invoice_for_print(invoice)
        if not mapped_invoice or not mapped_invoice.get('invoice_meta', {}).get('number'):
            QMessageBox.warning(parent, "Print Error", "No valid invoice selected. Cannot print.")
            return

        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_pdf:
                tmp_pdf_path = tmp_pdf.name
            generate_invoice_pdf(mapped_invoice, tmp_pdf_path)

            opened = False
            if sys.platform.startswith('win'):
                try:
                    os.startfile(tmp_pdf_path, 'print')
                    opened = True
                except OSError:
                    # No print association, try just opening
                    try:
                        os.startfile(tmp_pdf_path)
                        opened = True
                    except OSError:
                        pass
            elif sys.platform.startswith('darwin'):
                opened = os.system(f'open "{tmp_pdf_path}"') == 0
            else:
                opened = os.system(f'xdg-open "{tmp_pdf_path}"') == 0

            if not opened:
                QMessageBox.information(parent, "Print Info", f"Invoice PDF generated at:\n{tmp_pdf_path}\n\nNo application is associated to print or open PDF files. Please open and print it manually.")
        except Exception as e:
            log_error("Error printing invoice PDF", exception=e, logger_name='billing_errors')
            QMessageBox.critical(parent, "Print Error", f"Failed to print invoice: {e}")

    @staticmethod
    def export_to_excel(table: QTableWidget, report_name: str, parent_widget: QWidget):
        """
        Export the report table to an Excel (.xlsx) file.

        Args:
            table: QTableWidget containing report data
            report_name: Name of the report for filename
            parent_widget: Parent widget for dialogs
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            filename, _ = QFileDialog.getSaveFileName(
                parent_widget,
                f"Export {report_name} Report",
                f"{report_name.lower().replace(' ', '_')}_report.xlsx",
                "Excel Files (*.xlsx);;All Files (*.*)"
            )

            if not filename:
                return

            # Ensure .xlsx extension
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'

            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = report_name[:31]  # Excel sheet name limit

            # Define styles
            header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            header_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            data_font = Font(name='Calibri', size=11)
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

            # Write headers
            headers = []
            for col in range(table.columnCount()):
                header_item = table.horizontalHeaderItem(col)
                header_text = header_item.text() if header_item else f"Column {col + 1}"
                headers.append(header_text)

                cell = ws.cell(row=1, column=col + 1, value=header_text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Set header row height
            ws.row_dimensions[1].height = 30

            # Write data rows
            for row in range(table.rowCount()):
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    cell_value = item.text() if item else ''

                    # Try to convert to number for better Excel formatting
                    if cell_value.startswith(f"{get_currency_symbol()}"):
                        # Remove currency symbol and commas for numeric values
                        numeric_value = cell_value.replace(get_currency_symbol(), '').replace(',', '').strip()
                        try:
                            cell_value = float(numeric_value)
                        except ValueError:
                            pass  # Keep as string if conversion fails

                    cell = ws.cell(row=row + 2, column=col + 1, value=cell_value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border

                    # Format currency cells
                    if isinstance(cell_value, (int, float)):
                        cell.number_format = f'"{get_currency_symbol()}"#,##0.00'

            # Auto-adjust column widths with improved algorithm
            for col in range(1, table.columnCount() + 1):
                column_letter = get_column_letter(col)
                max_length = 0

                # Check header length
                header_length = len(str(headers[col - 1]))
                max_length = max(max_length, header_length)

                # Check ALL data rows for accurate sizing (not just sample)
                for row in range(2, ws.max_row + 1):
                    try:
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value:
                            cell_length = len(str(cell_value))
                            max_length = max(max_length, cell_length)
                    except:
                        pass

                # Set column width with appropriate limits
                # Minimum 12 characters, maximum 60 characters
                adjusted_width = min(max(max_length + 3, 12), 60)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Apply text wrapping to all data cells to prevent overflow
            from openpyxl.styles import Alignment as OpenpyxlAlignment
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.alignment:
                        cell.alignment = OpenpyxlAlignment(
                            horizontal=cell.alignment.horizontal,
                            vertical='center',
                            wrap_text=True
                        )

            # Freeze header row
            ws.freeze_panes = 'A2'

            # Save workbook
            wb.save(filename)

            log_info(f"Report exported to Excel: {filename}", 'billing_app')
            QMessageBox.information(
                parent_widget,
                "Success",
                f"Report exported successfully to Excel!\n\n{filename}"
            )

        except ImportError as e:
            log_error(f"openpyxl not installed", exception=e, logger_name='billing_errors')
            QMessageBox.critical(
                parent_widget,
                "Export Error",
                "Excel export requires 'openpyxl' package.\nPlease install it: pip install openpyxl"
            )
        except Exception as e:
            log_error(f"Failed to export report to Excel", exception=e, logger_name='billing_errors')
            QMessageBox.critical(
                parent_widget,
                "Export Error",
                f"Failed to export report to Excel:\n{str(e)}"
            )

    @staticmethod
    def export_to_pdf(table: QTableWidget, report_name: str, parent_widget: QWidget):
        """
        Export table data to PDF file with proper formatting.

        Args:
            table: QTableWidget containing report data
            report_name: Name of the report for filename
            parent_widget: Parent widget for dialogs
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.pdfgen import canvas
            
            filename, _ = QFileDialog.getSaveFileName(
                parent_widget,
                f"Export {report_name} Report",
                f"{report_name.lower().replace(' ', '_')}_report.pdf",
                "PDF Files (*.pdf);;All Files (*.*)"
            )
            
            if not filename:
                return
            
            # Ensure .pdf extension
            if not filename.endswith('.pdf'):
                filename += '.pdf'
            
            # Create PDF document (landscape for better table fit)
            doc = SimpleDocTemplate(
                filename,
                pagesize=landscape(A4),
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30
            )
            
            # Container for the 'Flowable' objects
            elements = []
            
            # Define styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#000000'),
                spaceAfter=12,
                alignment=1  # Center alignment
            )
            
            # Add title
            title = Paragraph(f"<b>{report_name} Report</b>", title_style)
            elements.append(title)
            
            # Add date
            from datetime import datetime
            date_str = datetime.now().strftime("%B %d, %Y %I:%M %p")
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#666666'),
                spaceAfter=20,
                alignment=1  # Center alignment
            )
            date_para = Paragraph(f"Generated on: {date_str}", date_style)
            elements.append(date_para)
            elements.append(Spacer(1, 0.2 * inch))
            
            # Prepare table data
            table_data = []
            
            # Headers
            headers = []
            for col in range(table.columnCount()):
                header_item = table.horizontalHeaderItem(col)
                header_text = header_item.text() if header_item else f"Column {col + 1}"
                headers.append(header_text)
            table_data.append(headers)
            
            # Data rows
            for row in range(table.rowCount()):
                row_data = []
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    cell_value = item.text() if item else ''
                    row_data.append(cell_value)
                table_data.append(row_data)
            
            # Create table with intelligent column width calculation
            if table_data:
                # Calculate column widths based on content length
                page_width = landscape(A4)[0] - 60  # Subtract margins
                num_cols = len(table_data[0])
                
                # Calculate max length for each column (including header)
                col_max_lengths = []
                for col_idx in range(num_cols):
                    max_len = len(str(headers[col_idx]))  # Start with header length
                    
                    # Check data rows (sample first 50 rows for performance)
                    for row_idx in range(1, min(51, len(table_data))):
                        if col_idx < len(table_data[row_idx]):
                            cell_len = len(str(table_data[row_idx][col_idx]))
                            max_len = max(max_len, cell_len)
                    
                    col_max_lengths.append(max_len)
                
                # Calculate proportional widths based on content
                total_chars = sum(col_max_lengths)
                col_widths = []
                
                if total_chars > 0:
                    for max_len in col_max_lengths:
                        # Proportional width with minimum of 0.5 inch
                        width = max((max_len / total_chars) * page_width, 0.5 * inch)
                        col_widths.append(width)
                    
                    # Adjust if total exceeds page width
                    total_width = sum(col_widths)
                    if total_width > page_width:
                        scale_factor = page_width / total_width
                        col_widths = [w * scale_factor for w in col_widths]
                else:
                    # Fallback to equal distribution
                    col_width = page_width / num_cols
                    col_widths = [col_width] * num_cols
                
                # Wrap long text in Paragraph objects to prevent overflow
                from reportlab.lib.styles import ParagraphStyle
                cell_style = ParagraphStyle(
                    'CellStyle',
                    parent=styles['Normal'],
                    fontSize=8,
                    leading=10,
                    wordWrap='CJK',
                    alignment=0  # Left alignment
                )
                
                # Convert long text cells to Paragraph objects
                formatted_table_data = []
                for row_idx, row in enumerate(table_data):
                    formatted_row = []
                    for col_idx, cell in enumerate(row):
                        cell_str = str(cell)
                        # Wrap long text (> 30 chars) in Paragraph for automatic wrapping
                        if len(cell_str) > 30 and row_idx > 0:  # Don't wrap headers
                            formatted_row.append(Paragraph(cell_str, cell_style))
                        else:
                            formatted_row.append(cell_str)
                    formatted_table_data.append(formatted_row)
                
                pdf_table = Table(formatted_table_data, colWidths=col_widths, repeatRows=1)
                
                # Apply table style with text wrapping support
                table_style = TableStyle([
                    # Header style
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#000000')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('TOPPADDING', (0, 0), (-1, 0), 12),
                    
                    # Data rows style
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('TOPPADDING', (0, 1), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                    
                    # Grid
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    
                    # Alternating row colors
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')])
                ])
                
                pdf_table.setStyle(table_style)
                elements.append(pdf_table)
            
            # Build PDF
            doc.build(elements)
            
            log_info(f"Report exported to PDF: {filename}", 'billing_app')
            QMessageBox.information(
                parent_widget,
                "Success",
                f"Report exported successfully to PDF!\n\n{filename}"
            )
        
        except ImportError as e:
            log_error(f"reportlab not installed", exception=e, logger_name='billing_errors')
            QMessageBox.critical(
                parent_widget,
                "Export Error",
                "PDF export requires 'reportlab' package.\nPlease install it: pip install reportlab"
            )
        except Exception as e:
            log_error(f"Failed to export report to PDF", exception=e, logger_name='billing_errors')
            QMessageBox.critical(
                parent_widget,
                "Export Error",
                f"Failed to export report to PDF:\n{str(e)}"
            )


class SummaryCardManager:
    """Manages summary card creation and updates."""
    
    @staticmethod
    def create_summary_cards(titles: List[str], colors: dict) -> QFrame:
        """
        Create summary cards with titles. Responsive: 3 per row, wrap as needed.
        Args:
            titles: List of card titles
            colors: Color scheme dictionary
        Returns:
            QFrame containing all summary cards
        """
        from PyQt6.QtWidgets import QGridLayout
        summary_frame = QFrame()
        summary_frame.setStyleSheet("""QFrame { background-color: #000000; border-radius: 8px; border: 2px solid #777777; padding: 15px; }""")
        summary_layout = QGridLayout(summary_frame)
        summary_layout.setContentsMargins(15, 15, 15, 15)
        summary_layout.setSpacing(20)
        cards_per_row = 3
        for i, title in enumerate(titles):
            card = SummaryCardManager._create_single_card(title, colors)
            row = i // cards_per_row
            col = i % cards_per_row
            summary_layout.addWidget(card, row, col)
        # Make columns stretch equally
        for col in range(cards_per_row):
            summary_layout.setColumnStretch(col, 1)
        return summary_frame
    
    @staticmethod
    def _create_single_card(title: str, colors: dict) -> QFrame:
        """Create a single summary card."""
        card = QFrame()
        card.setStyleSheet("""
            QFrame {
                background-color: #0F0F0F;
                border-radius: 8px;
                border: 2px solid #777777;
                padding: 15px;
            }
        """)
        card_layout = QVBoxLayout(card)
        card_layout.setSpacing(6)
        
        title_label = QLabel(title)
        title_label.setStyleSheet("""
            color: #FFFFFF;
            font-size: 14px;
            font-weight: bold;
        """)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        card_layout.addWidget(title_label)
        
        value_label = QLabel(format_currency(0))
        value_label.setProperty('summary_value', True)
        value_label.setStyleSheet("""
            color: #FFFFFF;
            font-size: 24px;
            font-weight: bold;
        """)
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        card_layout.addWidget(value_label)
        
        return card
    
    @staticmethod
    def update_summary_cards(frame: QFrame, values: List[str]):
        """
        Update summary card values. Works with both HBoxLayout and GridLayout.
        
        Args:
            frame: Summary cards container frame
            values: List of new values to display
        """
        from PyQt6.QtWidgets import QGridLayout
        
        layout = frame.layout()
        if not layout:
            return
        
        # Handle grid layout (used for 6 cards)
        if isinstance(layout, QGridLayout):
            value_index = 0
            for row in range(layout.rowCount()):
                for col in range(layout.columnCount()):
                    if value_index >= len(values):
                        return
                    item = layout.itemAtPosition(row, col)
                    if item and item.widget():
                        card = item.widget()
                        if isinstance(card, QFrame):
                            # Find the value label in this specific card
                            for label in card.findChildren(QLabel):
                                if label.property('summary_value'):
                                    label.setText(values[value_index])
                                    break
                    value_index += 1
        else:
            # Handle horizontal layout (original behavior)
            for i in range(min(layout.count(), len(values))):
                item = layout.itemAt(i)
                if item and item.widget():
                    card = item.widget()
                    if isinstance(card, QFrame):
                        # Find the value label in this specific card
                        for label in card.findChildren(QLabel):
                            if label.property('summary_value'):
                                label.setText(values[i])
                                break


def create_report_header(title: str, description: str, colors: dict) -> QWidget:
    """
    Create consistent styled header for reports.
    
    Args:
        title: Main title text with emoji
        description: Subtitle/description
        colors: Color scheme dictionary
        
    Returns:
        QWidget containing the styled header
    """
    header_widget = QWidget()
    header_layout = QVBoxLayout(header_widget)
    header_layout.setContentsMargins(0, 0, 0, 15)
    header_layout.setSpacing(5)
    
    # Main title
    title_label = QLabel(title)
    title_label.setStyleSheet(f"""
        QLabel {{
            color: {colors['accent_primary']};
            font-size: 26px;
            font-weight: bold;
            letter-spacing: 0.5px;
        }}
    """)
    header_layout.addWidget(title_label)
    
    # Description
    if description:
        desc_label = QLabel(description)
        desc_label.setStyleSheet(f"""
            QLabel {{
                color: {colors['text_secondary']};
                font-size: 14px;
                font-weight: 400;
                margin-top: 5px;
            }}
        """)
        header_layout.addWidget(desc_label)
    
    return header_widget


def show_no_records_message(parent_widget: QWidget, report_name: str):
    """
    Show informative message when no records match filters.
    (Dialog removed - now just logs the message)
    
    Args:
        parent_widget: Parent widget for dialog
        report_name: Name of the report
    """
    from travel_billing_software.utils.logger import log_info
    log_info(f"No records found for {report_name} with current filters", 'billing_app')
    # Dialog removed - table will show empty state instead
